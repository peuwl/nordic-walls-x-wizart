#!/usr/bin/env python3
"""
Nordic Walls → Wizart import pipeline (Electron app version).

Usage:
    python3 run_import.py <shopify_export.csv> <output_dir>

Outputs a zip to <output_dir>/ and prints:
    ZIP_OUTPUT: <absolute_path_to_zip>
"""

import csv
import os
import shutil
import sys
import warnings

# Force UTF-8 output on Windows (avoids cp1252 UnicodeEncodeError)
if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
from collections import OrderedDict
from datetime import date
from urllib.parse import urlparse

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
warnings.filterwarnings("ignore", message=".*NotOpenSSLWarning.*")
warnings.filterwarnings("ignore", message=".*urllib3.*")

import requests
from PIL import Image
import openpyxl

# ---------------------------------------------------------------------------
# Config — handle both normal execution and PyInstaller frozen bundle
# ---------------------------------------------------------------------------
if getattr(sys, "frozen", False):
    # Running as a PyInstaller --onefile binary; templates are extracted to _MEIPASS
    BUNDLE_DIR    = sys._MEIPASS
    TEMPLATES_DIR = os.path.join(BUNDLE_DIR, "templates")
else:
    SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
    TEMPLATES_DIR = os.path.join(SCRIPT_DIR, "..", "templates")
MURAL_TEMPLATE    = os.path.join(TEMPLATES_DIR, "Default mapping template_Wall_mural (1).xlsx")
WALLPAPER_TEMPLATE = os.path.join(TEMPLATES_DIR, "Default mapping template_Wallpaper (2).xlsx")
SHOP_BASE_URL = "https://nordicwalls.com/products"

IMAGE_SRC_COL    = "Image Src"
IMAGE_POS_COL    = "Image Position"
REPLICABLE_COL   = "Replicable Mural (product.metafields.custom.replicable_mural)"
TILE_SIZE_COL    = "Replicable Tile Size in CM (product.metafields.custom.replicable_cm_tile_size)"
PRODUCT_TYPE_COL = "Product type (product.metafields.custom.product_type)"
TAGS_COL         = "Tags"


# ---------------------------------------------------------------------------
# Utilities
# ---------------------------------------------------------------------------
def clean_url(url):
    return urlparse(url)._replace(query="", fragment="").geturl()


def get_filename(url):
    return os.path.basename(clean_url(url)) if url else ""


def classify_product(rows):
    """
    3-step classification:
      1. 'Product type' metafield -> Wallpaper / Mural
      2. 'Tags' column fallback   -> Patterns / Wall Murals / Murals
      3. Default                  -> mural (matches old 'other' bucket behaviour)
    """
    pt = next(
        (r.get(PRODUCT_TYPE_COL, "").strip() for r in rows if r.get(PRODUCT_TYPE_COL, "").strip()),
        ""
    ).lower()
    if pt == "wallpaper":
        return "pattern"
    if pt == "mural":
        return "mural"

    tags_raw = next(
        (r.get(TAGS_COL, "").strip() for r in rows if r.get(TAGS_COL, "").strip()),
        ""
    )
    tags_lower = {t.strip().lower() for t in tags_raw.split(",") if t.strip()}
    if "patterns" in tags_lower or "pattern" in tags_lower:
        return "pattern"
    if "wall murals" in tags_lower or "murals" in tags_lower or "mural" in tags_lower:
        return "mural"

    return "other"


def get_last_image(rows):
    image_rows = [(i, r) for i, r in enumerate(rows) if r[IMAGE_SRC_COL].strip()]
    if not image_rows:
        return None, None, None
    def pos_key(item):
        try:
            return int(item[1][IMAGE_POS_COL])
        except (ValueError, KeyError):
            return item[0]
    last_idx, last_row = sorted(image_rows, key=pos_key)[-1]
    return last_idx, last_row[IMAGE_SRC_COL].strip(), get_filename(last_row[IMAGE_SRC_COL].strip())


def download_image(url, dest_path):
    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        with open(dest_path, "wb") as f:
            f.write(resp.content)
        return True
    except Exception as e:
        print(f"  ERROR downloading {url}: {e}", flush=True)
        return False


def is_portrait(path):
    try:
        with Image.open(path) as img:
            w, h = img.size
            return h > w
    except Exception as e:
        print(f"  ERROR reading {path}: {e}", flush=True)
        return False


def parse_title(title):
    title = title.strip()
    if " | " in title:
        parts = title.split(" | ", 1)
        collection = parts[0].strip()
        color = parts[1].strip()
        return collection, f"{collection} {color}", color
    return title, title, ""


# ---------------------------------------------------------------------------
# Step 1: Read & split CSV
# ---------------------------------------------------------------------------
def read_and_split(input_csv, mode="auto"):
    """
    mode: 'auto'        — classify each product via classify_product()
          'wallpapers'  — treat all products as wallpapers
          'murals'      — treat all products as murals
    Returns (products, fieldnames, pattern_handles, mural_handles, other_handles).
    """
    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        fieldnames = reader.fieldnames
        all_rows = list(reader)

    products = OrderedDict()
    for row in all_rows:
        products.setdefault(row["Handle"], []).append(row)

    pattern_handles, mural_handles, other_handles = [], [], []
    for handle, rows in products.items():
        if mode == "wallpapers":
            pattern_handles.append(handle)
            continue
        if mode == "murals":
            mural_handles.append(handle)
            continue

        kind = classify_product(rows)
        if kind == "pattern":
            pattern_handles.append(handle)
        elif kind == "mural":
            mural_handles.append(handle)
        else:
            other_handles.append(handle)

    return products, fieldnames, pattern_handles, mural_handles, other_handles


def build_rows_single_image(rows):
    last_idx, _, _ = get_last_image(rows)
    out = []
    for i, row in enumerate(rows):
        new_row = dict(row)
        if i != last_idx and new_row[IMAGE_SRC_COL].strip():
            new_row[IMAGE_SRC_COL] = ""
            new_row[IMAGE_POS_COL] = ""
        elif i == last_idx:
            new_row[IMAGE_POS_COL] = "1"
        out.append(new_row)
    return out


def write_csv(path, handles, products, fieldnames, single_image=True):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for handle in handles:
            rows = build_rows_single_image(products[handle]) if single_image else products[handle]
            writer.writerows(rows)


# ---------------------------------------------------------------------------
# Step 2: Download images
# ---------------------------------------------------------------------------
def download_images(handles, products, dest_dir, check_portrait=False, portrait_dir=None):
    os.makedirs(dest_dir, exist_ok=True)
    portrait_list = []
    failures = []
    for handle in handles:
        _, url, filename = get_last_image(products[handle])
        if not url:
            print(f"  [no image] {handle}", flush=True)
            continue
        dest = os.path.join(dest_dir, filename)
        if os.path.exists(dest):
            print(f"  [skip] {filename}", flush=True)
        else:
            print(f"  Downloading {filename} ...", flush=True)
            if not download_image(url, dest):
                failures.append(url)
                continue
        if check_portrait and portrait_dir and os.path.exists(dest) and is_portrait(dest):
            os.makedirs(portrait_dir, exist_ok=True)
            portrait_dest = os.path.join(portrait_dir, filename)
            shutil.copy2(dest, portrait_dest)
            portrait_list.append((handle, filename))
    return portrait_list, failures


# ---------------------------------------------------------------------------
# Step 3: Generate Wizart xlsx
# ---------------------------------------------------------------------------
def read_products_for_xlsx(handles, products):
    result = OrderedDict()
    for handle in handles:
        rows = products[handle]
        title = next((r["Title"].strip() for r in rows if r["Title"].strip()), handle)
        _, url, filename = get_last_image(rows)
        replicable = any(
            r.get(REPLICABLE_COL, "").strip().upper() == "TRUE" for r in rows
        )
        tile_raw = next(
            (r.get(TILE_SIZE_COL, "").strip() for r in rows if r.get(TILE_SIZE_COL, "").strip()),
            ""
        )
        tile_raw_clean = tile_raw.lstrip("'").strip()
        tile_cm = int(float(tile_raw_clean)) if tile_raw_clean else None
        result[handle] = {
            "title": title,
            "image": filename,
            "replicable": replicable,
            "tile_size_cm": tile_cm,
        }
    return result


def get_header_map(ws):
    return {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}


def generate_mural_xlsx(handle_groups, products, output_path):
    all_handles = []
    for group in handle_groups:
        all_handles.extend(group)
    pdata = read_products_for_xlsx(all_handles, products)

    wb = openpyxl.load_workbook(MURAL_TEMPLATE)
    ws = wb.active
    hdr = get_header_map(ws)

    row_idx = 2
    for handle, p in pdata.items():
        if not p["image"]:
            print(f"  [skip, no image] {handle}", flush=True)
            continue
        collection, product_name, color = parse_title(p["title"])

        def w(field, value):
            if field in hdr:
                ws.cell(row=row_idx, column=hdr[field], value=value)

        w("brand_name", "Nordic Walls")
        w("collection_name", collection)
        w("product_name", product_name)
        w("unique_sku_id", handle)
        w("product_image", p["image"])
        w("product_width", 1.5)
        w("product_length", 1.0)
        w("pattern_length", 1.0)
        w("pattern_width", 1.5)
        w("repeatable", "yes" if p["replicable"] else "no")
        w("customized_size", "yes")
        w("color", color)
        w("product_link", f"{SHOP_BASE_URL}/{handle}")
        w("application_surface", "wall")
        w("material", "Wallpaper")
        w("product_availability", "in_stock")
        w("product_regular_price", 329)
        w("context_currency", "EUR")
        row_idx += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    count = row_idx - 2
    print(f"  Wrote {count} murals → wizart_wall_murals.xlsx", flush=True)
    return count


def generate_pattern_xlsx(pattern_handles, products, output_path):
    pdata = read_products_for_xlsx(pattern_handles, products)

    wb = openpyxl.load_workbook(WALLPAPER_TEMPLATE)
    ws = wb.active
    hdr = get_header_map(ws)

    row_idx = 2
    for handle, p in pdata.items():
        if not p["image"]:
            print(f"  [skip, no image] {handle}", flush=True)
            continue
        collection, product_name, color = parse_title(p["title"])
        if p["tile_size_cm"]:
            tile_m = round(p["tile_size_cm"] / 100, 4)
        else:
            tile_m = 0.5
            print(f"  WARNING: no tile size for '{handle}', defaulting to 0.5m", flush=True)

        def w(field, value):
            if field in hdr:
                ws.cell(row=row_idx, column=hdr[field], value=value)

        w("brand_name", "Nordic Walls")
        w("collection_name", collection)
        w("product_name", product_name)
        w("unique_sku_id", handle)
        w("product_image", p["image"])
        w("product_width", 1.5)
        w("repeat_width", tile_m)
        w("pattern_repeat", tile_m)
        w("pattern_offset", 0)
        w("color", color)
        w("product_link", f"{SHOP_BASE_URL}/{handle}")
        w("mural", "no")
        w("product_availability", "in_stock")
        w("product_regular_price", 329)
        w("context_currency", "EUR")
        row_idx += 1

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    count = row_idx - 2
    print(f"  Wrote {count} patterns → wizart_wallpapers.xlsx", flush=True)
    return count


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def parse_args(argv):
    """Parse positional CSV + output_dir args plus optional --mode <auto|wallpapers|murals>."""
    mode = "auto"
    positional = []
    i = 1
    while i < len(argv):
        if argv[i] == "--mode" and i + 1 < len(argv):
            mode = argv[i + 1].lower()
            i += 2
            continue
        positional.append(argv[i])
        i += 1
    if mode not in ("auto", "wallpapers", "murals"):
        print(f"ERROR: invalid --mode '{mode}' (expected auto|wallpapers|murals)")
        sys.exit(1)
    if len(positional) < 2:
        print("Usage: python3 run_import.py <shopify_export.csv> <output_dir> [--mode auto|wallpapers|murals]")
        sys.exit(1)
    return positional[0], positional[1], mode


def main():
    input_csv, output_dir, mode = parse_args(sys.argv)

    if not os.path.exists(input_csv):
        print(f"ERROR: File not found: {input_csv}")
        sys.exit(1)

    do_murals     = mode in ("auto", "murals")
    do_wallpapers = mode in ("auto", "wallpapers")
    mode_suffix   = {"auto": "", "wallpapers": " (Wallpapers)", "murals": " (Murals)"}[mode]

    now = __import__("datetime").datetime.now()
    folder_name = f"Wizart Import Files{mode_suffix} - {now.strftime('%Y-%m-%d %H.%M')}"
    out_dir  = os.path.join(output_dir, folder_name)
    misc_dir = os.path.join(out_dir, "~misc")
    os.makedirs(out_dir, exist_ok=True)
    os.makedirs(misc_dir, exist_ok=True)
    print(f"Mode: {mode}", flush=True)
    print(f"Output folder: {out_dir}/\n", flush=True)

    pattern_csv  = os.path.join(misc_dir, "products_pattern.csv")
    mural_csv    = os.path.join(misc_dir, "products_mural.csv")
    portrait_csv = os.path.join(misc_dir, "products_mural_portrait.csv")
    other_csv    = os.path.join(misc_dir, "products_other.csv")
    pattern_dir  = os.path.join(misc_dir, "pattern_images")
    mural_dir    = os.path.join(misc_dir, "mural_images")

    mural_xlsx   = os.path.join(out_dir, "wizart_wall_murals.xlsx")
    pattern_xlsx = os.path.join(out_dir, "wizart_wallpapers.xlsx")

    # --- Step 1: Split CSV ---
    print("── Step 1: Splitting CSV ──────────────────────────────", flush=True)
    products, fieldnames, pattern_handles, mural_handles, other_handles = read_and_split(input_csv, mode=mode)
    if do_wallpapers:
        write_csv(pattern_csv, pattern_handles, products, fieldnames)
        print(f"  Pattern products : {len(pattern_handles)}", flush=True)
    if do_murals:
        write_csv(mural_csv, mural_handles, products, fieldnames)
        print(f"  Mural products   : {len(mural_handles)}", flush=True)
    if mode == "auto":
        write_csv(other_csv, other_handles, products, fieldnames, single_image=False)
        print(f"  Other products   : {len(other_handles)}", flush=True)

    # --- Step 2: Download images ---
    print("\n── Step 2: Downloading images ─────────────────────────", flush=True)
    pattern_failures, mural_failures, other_failures = [], [], []
    portrait_handles, landscape_handles = [], []

    if do_wallpapers:
        print(f"Pattern images ({len(pattern_handles)}):", flush=True)
        _, pattern_failures = download_images(pattern_handles, products, pattern_dir)

    if do_murals:
        print(f"\nMural images ({len(mural_handles)}):", flush=True)
        portrait_pairs, mural_failures = download_images(
            mural_handles, products, mural_dir, check_portrait=True, portrait_dir=None
        )
        portrait_handles  = [h for h, _ in portrait_pairs]
        landscape_handles = [h for h in mural_handles if h not in portrait_handles]

        if mode == "auto" and other_handles:
            print(f"\nOther images ({len(other_handles)}):", flush=True)
            _, other_failures = download_images(
                [h for h in other_handles if get_last_image(products[h])[1]],
                products, mural_dir
            )

        print(f"\n  Landscape murals : {len(landscape_handles)}", flush=True)
        print(f"  Portrait murals  : {len(portrait_handles)}", flush=True)
        write_csv(portrait_csv, portrait_handles, products, fieldnames)

    # --- Step 3: Generate Wizart xlsx ---
    print("\n── Step 3: Generating Wizart import files ─────────────", flush=True)
    n_murals = n_patterns = 0
    if do_murals:
        groups = [landscape_handles, portrait_handles]
        if mode == "auto":
            groups.append(other_handles)
        n_murals = generate_mural_xlsx(groups, products, mural_xlsx)
    if do_wallpapers:
        n_patterns = generate_pattern_xlsx(pattern_handles, products, pattern_xlsx)

    # --- Step 4: Zip image folders ---
    print("\n── Step 4: Zipping image folders ──────────────────────", flush=True)
    if do_murals:
        shutil.make_archive(os.path.join(out_dir, "wall_murals_images"), "zip", misc_dir, "mural_images")
        print("  wall_murals_images.zip created", flush=True)
    if do_wallpapers:
        shutil.make_archive(os.path.join(out_dir, "wallpaper_images"), "zip", misc_dir, "pattern_images")
        print("  wallpaper_images.zip created", flush=True)

    # --- Step 5: Zip entire folder ---
    print("\n── Step 5: Zipping output folder ──────────────────────", flush=True)
    zip_base = os.path.join(output_dir, folder_name)
    shutil.make_archive(zip_base, "zip", output_dir, folder_name)
    shutil.rmtree(out_dir)
    zip_path = zip_base + ".zip"
    print(f"  {folder_name}.zip created", flush=True)

    # --- Summary ---
    all_failures = pattern_failures + mural_failures + other_failures
    print(f"\n── Done ────────────────────────────────────────────────", flush=True)
    print(f"  Total products : {n_murals + n_patterns}", flush=True)
    if do_murals:
        print(f"  Murals xlsx    : {n_murals} products", flush=True)
    if do_wallpapers:
        print(f"  Patterns xlsx  : {n_patterns} products", flush=True)
    if all_failures:
        print(f"\n  Failed downloads ({len(all_failures)}):", flush=True)
        for u in all_failures:
            print(f"    {u}", flush=True)
    else:
        print("  All downloads succeeded.", flush=True)

    # Signal the zip path to the Electron main process
    print(f"\nZIP_OUTPUT:{zip_path}", flush=True)


if __name__ == "__main__":
    main()
